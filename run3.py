import nltk
import re
import requests
import json
import xlsxwriter
import emoji
import openpyxl
from nltk.tag.stanford import StanfordTagger as POS_tag

#importing stanford part of speech tagger
_path_to_model='stanford-postagger/models/english-bidirectional-distsim.tagger'
_path_to_jar='stanford-postagger/stanford-postagger.jar'
st=POS_tag(_path_to_model,_path_to_jar)

#creating excel sheet to populate results

# wk=xlsxwriter.Workbook('dataset.xlsx')
# worksheet=wk.add_worksheet('Sheet1')
# worksheet.cell(row=0,0,'Original Message')
# worksheet.cell(row=0,1,'Has Emoticon?')
# worksheet.cell(row=0,2,'Contains URL?')
# worksheet.cell(row=0,3,"No. of characters")
# worksheet.cell(row=0,4,'No. of Uppercase Characters')
# worksheet.cell(row=0,5,'Upper Case Score')
# worksheet.cell(row=0,6,'Cleaned SMS')
# worksheet.cell(row=0,7,'No. of Characters in Cleaned SMS')
# worksheet.cell(row=0,8,'Cleaned Score')
# worksheet.cell(row=0,9,'Total Score')
# worksheet.cell(row=0,10,'Model Prediction')
# worksheet.cell(row=0,11,'Naive Bayes Prediction')
# wk.close()

filename = 'dataset-r3.xlsx'
wk = openpyxl.Workbook()
worksheet = wk.get_sheet_by_name('Sheet')
worksheet.cell(row=1,column=1).value = 'Original Message'
worksheet.cell(row=1,column=2).value ='Has Emoticon?'
worksheet.cell(row=1,column=3).value ='Contains URL?'
worksheet.cell(row=1,column=4).value ="No. of characters"
worksheet.cell(row=1,column=5).value ='No. of Uppercase Characters'
worksheet.cell(row=1,column=6).value ='Upper Case Score'
worksheet.cell(row=1,column=7).value ='Cleaned SMS'
worksheet.cell(row=1,column=8).value ='No. of Characters in Cleaned SMS'
worksheet.cell(row=1,column=9).value ='Cleaned Score'
worksheet.cell(row=1,column=10).value ='Total Score'
worksheet.cell(row=1,column=11).value ='Model Prediction'
worksheet.cell(row=1,column=12).value ='Naive Bayes Prediction'
wk.save(filename)
l=0
threshhold=0.4
with open('smsdata.txt','r') as f:
    for sms in f:
        l=l+1
        if(l<16346):
            continue
        # wk = openpyxl.load_workbook(filename) 
        # worksheet = wk.get_sheet_by_name('Sheet')
        worksheet.cell(row=l,column=1).value =sms.decode('utf-8')
        has_emoji=False
        for c in sms:
            if c in emoji.UNICODE_EMOJI:
                has_emoji=True
                break
        worksheet.cell(row=l,column=2).value=str(has_emoji)
        if has_emoji==True:
            worksheet.cell(row=l,column=3).value ='N/A'
            worksheet.cell(row=l,column=4).value ='N/A'
            worksheet.cell(row=l,column=5).value ='N/A'
            worksheet.cell(row=l,column=6).value ='N/A'
            worksheet.cell(row=l,column=7).value ='N/A'
            worksheet.cell(row=l,column=8).value ='N/A'
            worksheet.cell(row=l,column=9).value ='N/A'
            worksheet.cell(row=l,column=10).value ='N/A'
            worksheet.cell(row=1,column=11).value ='HAM'
        else:
            spam_score=0
            url_regex = r"(http:\/\/www\.|https:\/\/www\.|http:\/\/|https:\/\/)?[a-z0-9]+([\-\.]{1}[a-z0-9]+)*\.[a-z]{2,5}(:[0-9]{1,5})?(\/.*)?"
            has_url= len(re.findall(url_regex, sms)) > 0
            print (l, re.findall(url_regex, sms))
            if has_url==True:
                worksheet.cell(row=l,column=3).value='Yes'
                spam_score+=0.1
            else:
                worksheet.cell(row=l,column=3).value ='No'
            
            length_sms=len(sms)-sms.count(' ')
            worksheet.cell(row=l,column=4).value=length_sms
            
            upper_chars=sum(1 for c in sms if c.isupper())
            worksheet.cell(row=l,column=5).value=upper_chars
            
            upper_case_score=(upper_chars/float(length_sms))*0.2
            worksheet.cell(row=l,column=6).value=str(upper_case_score)
            spam_score=spam_score+upper_case_score
            
            arr = ' '.join(w for w in re.split(r'\W+', sms.lower()))
            ar = re.sub(r'\w*\d\w*', '', arr).strip()
            tag_sms = nltk.pos_tag(ar.split())
            edit_sms = [word for word, tag in tag_sms if tag != 'NNP' and tag != 'NNPS' and tag != 'NN' and tag != 'FW']
            cleaned = ' '.join(edit_sms)
            worksheet.cell(row=l,column=7).value=cleaned.decode('utf-8')
            
            cleaned_sms_len=len(cleaned)-cleaned.count(' ')
            cleaned_sms_score=(1-((length_sms-cleaned_sms_len)/float(length_sms)))*0.7
            
            worksheet.cell(row=l,column=8).value=cleaned_sms_len
            
            worksheet.cell(row=l,column=9).value=cleaned_sms_score
            spam_score=spam_score+cleaned_sms_score
            
            worksheet.cell(row=l,column=10).value=spam_score
            
            if spam_score>=threshhold:
                worksheet.cell(row=l,column=11).value='SPAM'
            else:
                worksheet.cell(row=l,column=11).value='HAM'
            
            # payload = {"user": "sahil", "texts": [{"id": str(l), "textMessage": sms}]}
            # r=requests.post('https://glacial-hamlet-87000.herokuapp.com/mobileapp/', json=payload)
            # j = json.loads(r.text)
            # worksheet.cell(row=l,column=12).value=j['text']['cat']

        # wk.save(filename)
wk.save(filename)