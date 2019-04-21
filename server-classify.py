import requests
import json
import openpyxl

import os
os.environ['NO_PROXY'] = '127.0.0.1'
url = 'http://127.0.0.1:1504/classify/'

l=0
wk = openpyxl.Workbook()
worksheet = wk.get_sheet_by_name('Sheet')
payload = {"user": "sahil", "texts": []}
i=1
with open('smsdata.txt','r') as f:
    for sms in f:
        l+=1
        payload["texts"].append({"id": str(l), "textMessage": sms})
        i+=1
        if(i>400):
            r=requests.post(url, json=payload)
            body = json.loads(r.content)
            for t in body["texts"]:
                print (int(t["id"]), str(t["cat"]))
                worksheet.cell(row=int(t["id"]),column=12).value= str(t['cat'])
            i=0
            payload = {"user": "sahil", "texts": []}
            wk.save("dataset2.xlsx")

r=requests.post(url, json=payload)
body = json.loads(r.content)
for t in body["texts"]:
    print (int(t["id"]), str(t["cat"]))
    worksheet.cell(row=int(t["id"]),column=12).value= str(t['cat'])
i=0         
wk.save("dataset2.xlsx")